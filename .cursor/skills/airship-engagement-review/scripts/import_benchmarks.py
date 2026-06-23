#!/usr/bin/env python3
"""Import an Airship UA Benchmarks .xlsx into benchmarks.json (+ benchmarks.md).

Usage:
    python scripts/import_benchmarks.py <benchmarks.xlsx> [quarter_label]

The Airship "UA Benchmarks" workbook ships these sheets (Q1 2026 layout):
  - Push Opt-in Rate            : Quarter | DeviceFamily | Vertical | p10 | p50 | p90
  - Engagement Direct & Infl Opens : Quarter | DeviceFamily | Vertical | dir p10/p50/p90 | infl p10/p50/p90
  - Push Sends Per User- Month  : Quarter | DeviceFamily | Vertical | p10 | p50 | p90
  - Message Center Read Rate    : Quarter | Vertical | p10 | p50 | p90   (no device family)

Percentiles map to the workbook's Overview legend: Low=10th, Medium=50th, High=90th.
Re-run each quarter with the new file; this regenerates both resource files.
"""
import json
import os
import re
import sys
import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)

# Aliases let the skill match a client's industry to an Airship vertical.
VERTICAL_ALIASES = {
    "all_verticals": ["all", "overall", "cross-industry", "baseline", "global"],
    "business": ["b2b", "professional", "enterprise", "saas"],
    "charities_foundations_and_non_profit": ["non-profit", "nonprofit", "charity", "foundation", "ngo"],
    "education": ["edtech", "school", "university", "learning", "e-learning"],
    "entertainment": ["streaming", "ott", "vod", "music", "video", "movies"],
    "finance_insurance": ["bank", "banking", "insurance", "assurance", "fintech", "finance", "banque"],
    "food_drink": ["qsr", "restaurant", "food", "drink", "delivery", "fast food"],
    "gambling_gaming": ["gaming", "games", "casino", "betting", "gambling", "igaming"],
    "government": ["public sector", "gov", "administration", "civic"],
    "media": ["news", "publishing", "broadcaster", "press", "journalism", "tv", "magazine"],
    "medical_health_fitness": ["health", "fitness", "medical", "wellness", "healthcare", "pharma"],
    "retail": ["ecommerce", "e-commerce", "grocery", "distribution", "fashion", "shopping", "fmcg", "supermarket", "commerce", "marketplace"],
    "social": ["social network", "dating", "community", "messaging"],
    "sports_recreation": ["sports", "recreation", "outdoor", "league"],
    "travel_transportation": ["travel", "transport", "airline", "hospitality", "mobility", "hotel", "transportation"],
    "utility_productivity": ["utility", "productivity", "tools", "telecom", "telco", "operator", "carrier"],
}


def slug(s):
    return re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower()).strip("_")


def num(v):
    if v is None or v == "N/A":
        return None
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if vals[0] is None:
            continue
        rows.append(vals)
    return rows


def ensure(verticals, label):
    key = slug(label)
    if key not in verticals:
        verticals[key] = {"label": label, "aliases": VERTICAL_ALIASES.get(key, []), "metrics": {}}
    return verticals[key]


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    xlsx = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    verticals = {}
    quarter = sys.argv[2] if len(sys.argv) > 2 else None

    # Opt-in rate: Quarter | DeviceFamily | Vertical | p10 | p50 | p90
    for q, dev, vert, p10, p50, p90 in [r[:6] for r in sheet_rows(wb, "Push Opt-in Rate")]:
        quarter = quarter or q
        v = ensure(verticals, vert)
        v["metrics"].setdefault("optin_rate", {})[str(dev).lower()] = {"p10": num(p10), "p50": num(p50), "p90": num(p90)}

    # Direct & Influenced opens
    for row in sheet_rows(wb, "Engagement Direct & Infl Opens"):
        q, dev, vert = row[0], row[1], row[2]
        dp10, dp50, dp90, ip10, ip50, ip90 = row[3:9]
        v = ensure(verticals, vert)
        v["metrics"].setdefault("direct_open_rate", {})[str(dev).lower()] = {"p10": num(dp10), "p50": num(dp50), "p90": num(dp90)}
        if num(ip50) is not None:
            v["metrics"].setdefault("influenced_open_rate", {})[str(dev).lower()] = {"p10": num(ip10), "p50": num(ip50), "p90": num(ip90)}

    # Sends per user per month
    for q, dev, vert, p10, p50, p90 in [r[:6] for r in sheet_rows(wb, "Push Sends Per User- Month")]:
        v = ensure(verticals, vert)
        v["metrics"].setdefault("sends_per_user_month", {})[str(dev).lower()] = {"p10": num(p10), "p50": num(p50), "p90": num(p90)}

    # Message Center / Rich page read rate (no device family)
    mc_sheet = next((s for s in wb.sheetnames if "Read Rate" in s), None)
    if mc_sheet:
        for row in sheet_rows(wb, mc_sheet):
            q, vert, p10, p50, p90 = row[0], row[1], row[2], row[3], row[4]
            v = ensure(verticals, vert)
            v["metrics"].setdefault("message_center_read_rate", {})["all"] = {"p10": num(p10), "p50": num(p50), "p90": num(p90)}

    out = {
        "meta": {
            "source": "Airship User Engagement (UA) Benchmarks",
            "file": os.path.basename(xlsx),
            "published": quarter,
            "region": "global",
            "percentiles": {"low": "p10 (10th)", "median": "p50 (50th)", "high": "p90 (90th)"},
            "imported": datetime.date.today().isoformat(),
            "notes": (
                "Percentiles per Airship Overview legend: Low=10th, Medium=50th (median), High=90th. "
                "Opt-in/open/sends split by device family (ios/android/web); message_center_read_rate is vertical-only. "
                "sends_per_user_month is PER MONTH (multiply a weekly client pressure by ~4.33 to compare). "
                "Global sample; no region/locale split."
            ),
        },
        "metric_keys": {
            "optin_rate": {"label": "Push opt-in rate", "unit": "pct", "by": "device_family"},
            "direct_open_rate": {"label": "Direct open rate (directly tapped push / sends)", "unit": "pct", "by": "device_family"},
            "influenced_open_rate": {"label": "Influenced open rate (app opened <=12h of push / sends)", "unit": "pct", "by": "device_family"},
            "sends_per_user_month": {"label": "Push sends per user per MONTH", "unit": "sends_per_user_per_month", "by": "device_family"},
            "message_center_read_rate": {"label": "Message Center / rich page read rate", "unit": "pct", "by": "vertical_only"},
        },
        "verticals": dict(sorted(verticals.items())),
    }

    json_path = os.path.join(SKILL_DIR, "benchmarks.json")
    with open(json_path, "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"Wrote {json_path}: {len(out['verticals'])} verticals, quarter {quarter}")

    write_markdown(out)


def fmt_pct(d):
    if not d or d.get("p50") is None:
        return "—"
    return f"{d['p50']*100:.1f}% [{d['p10']*100:.1f}–{d['p90']*100:.1f}]"


def fmt_num(d):
    if not d or d.get("p50") is None:
        return "—"
    return f"{d['p50']:.1f} [{d['p10']:.1f}–{d['p90']:.1f}]"


def write_markdown(out):
    m = out["meta"]
    lines = [
        "# Industry benchmarks (engagement review)",
        "",
        f"Source: **{m['source']}** · file `{m['file']}` · published **{m['published']}** · "
        f"region **{m['region']}** · imported {m['imported']}.",
        "",
        "Machine-readable copy: `benchmarks.json` (scripts read that). Regenerate with",
        "`python scripts/import_benchmarks.py <file.xlsx>` when a new quarter arrives.",
        "",
        "## How the skill uses this file",
        "1. Determine the client's **industry** (confirm with the user; default = deduced from",
        "   brand research) and match it to a vertical below via its `aliases`.",
        "2. Show the client KPI next to the vertical **median (p50)** and the **[p10–p90] range**,",
        "   tagged `[Data]`, with the gap (points or ×). Percentiles: Low=10th, Medium=50th, High=90th.",
        "3. **Align definition + device family + denominator**: opt-in↔opt-in, per-platform↔per-platform.",
        "   `sends_per_user_month` is **per MONTH** — multiply a weekly client pressure by ~4.33 to compare.",
        "4. Always cite source + quarter + region. **No region/locale split** (global sample).",
        "5. **Confidence capped at Medium** (external/contextual). No matching vertical / metric →",
        "   state \"industry benchmark not available\"; **never fabricate**.",
        "",
        f"> Notes: {m['notes']}",
        "",
        "## Metric keys",
        "| Key | Meaning | Split by |",
        "|---|---|---|",
    ]
    for k, meta in out["metric_keys"].items():
        lines.append(f"| `{k}` | {meta['label']} | {meta['by']} |")
    lines += ["", "## Verticals & values", ""]
    for key, v in out["verticals"].items():
        lines.append(f"### {v['label']}  (`{key}`)")
        if v["aliases"]:
            lines.append(f"_aliases: {', '.join(v['aliases'])}_")
        met = v["metrics"]
        lines.append("")
        lines.append("| Metric | iOS | Android | Web/All |")
        lines.append("|---|---|---|---|")
        def cell(metric, fam, pct=True):
            d = met.get(metric, {}).get(fam)
            return (fmt_pct(d) if pct else fmt_num(d))
        lines.append(f"| Opt-in rate | {cell('optin_rate','ios')} | {cell('optin_rate','android')} | — |")
        lines.append(f"| Direct open rate | {cell('direct_open_rate','ios')} | {cell('direct_open_rate','android')} | {cell('direct_open_rate','web')} |")
        lines.append(f"| Influenced open rate | {cell('influenced_open_rate','ios')} | {cell('influenced_open_rate','android')} | — |")
        lines.append(f"| Sends/user/month | {cell('sends_per_user_month','ios',False)} | {cell('sends_per_user_month','android',False)} | — |")
        mc = met.get("message_center_read_rate", {}).get("all")
        lines.append(f"| Msg Center read rate | — | — | {fmt_pct(mc)} |")
        lines.append("")
    md_path = os.path.join(SKILL_DIR, "benchmarks.md")
    with open(md_path, "w") as f:
        f.write("\n".join(lines) + "\n")
    print(f"Wrote {md_path}")


if __name__ == "__main__":
    main()
